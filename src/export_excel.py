import os
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill, numbers, Font, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl import formatting as op_fm
from openpyxl.utils.exceptions import InvalidFileException, ReadOnlyWorkbookException

class ExportExcel:
    def __init__(self, output_file_data, output_file, info_text, language):
        self.output_file_data = output_file_data
        self.output_file = output_file
        self.info_text = info_text
        self.language = language

    def num_format(self, cell):
        cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        cell.alignment = Alignment(horizontal='right', vertical='center', wrapText=False)

    def write_to_excel(self):
        try:
            with pd.ExcelWriter(f'{self.output_file}', engine='openpyxl') as writer:
                for name, df, limits, oneport in self.output_file_data:
                    name = os.path.splitext(name)[0]
                    if oneport:
                        df["Injection"] = None
                        df["Error Injection"] = None
                        df.loc[:, 'Injection'] = pd.NA
                        df.loc[:, 'Error Injection'] = pd.NA

                    # Max Char length for sheets in excel = 31
                    df.to_excel(writer, sheet_name=name[:31], index=False)

                    wb = writer.book
                    ws = wb[name[:31]]

                    # Define Limits
                    self.infusion_upper_limit = limits[0][1]
                    self.infusion_lower_limit = limits[0][0]

                    self.injection_upper_limit = limits[1][1] if not oneport else None
                    self.injection_lower_limit = limits[1][0] if not oneport else None

                    self.set_column_widths(ws, oneport)
                    self.add_chart(ws, df, name, oneport)
                    self.add_conditional_formatting(ws, df, oneport)
                    self.add_data_to_sheet(ws, df, oneport)
        except PermissionError:
            self.info_text.insertPlainText(f'Error: File "{self.output_file}" is read-only. Please close this file and try again.\n')
        except Exception as e:
            self.info_text.insertPlainText(f'Error: {e}\n')

    def set_column_widths(self, ws, oneport):
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15 if not oneport else 5
        ws.column_dimensions.group(start='C', end='D', hidden=True)
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20 if not oneport else 5
        ws.column_dimensions['I'].width = 20

    def add_chart(self, ws, df, name, oneport):
        chart = BarChart()
        data = Reference(ws, min_col=1, max_col=2 if not oneport else 1, min_row=1, max_row=df.shape[0]+1)
        chart.add_data(data, titles_from_data=True)
        chart.title = name
        chart.x_axis.title = 'Cycle'
        chart.y_axis.title = 'Depth [mm]'
        chart.style = 2
        ws.add_chart(chart, 'I1')

    def add_conditional_formatting(self, ws, df, oneport):

        infusion_rule = op_fm.rule.CellIsRule(
            operator='lessThanOrEqual',
            formula=['$F$2'],
            stopIfTrue=False,
            font=Font(bold=True, color='E67E17'),
        )

        injection_rule = op_fm.rule.CellIsRule(
            operator='lessThanOrEqual',
            formula=['$G$2'],
            stopIfTrue=False,
            font=Font(bold=True, color="E67E17"),
        )

        ws.conditional_formatting.add(f'A2:A{df.shape[0]+1}', infusion_rule)
        ws.conditional_formatting.add(f'B2:B{df.shape[0]+1}', injection_rule) if not oneport else None

        # Min Value
        dxf = DifferentialStyle(fill=PatternFill(bgColor='B8F589'))
        rule = Rule(type='top10', rank=1, dxf=dxf)
        ws.conditional_formatting.add(f'A1:A{df.shape[0]+1}', rule)
        ws.conditional_formatting.add(f'B1:B{df.shape[0]+1}', rule) if not oneport else None

        # Max Value
        dxf = DifferentialStyle(fill=PatternFill(bgColor='FFC7CE'))
        rule = Rule(type='top10', bottom=True, rank=1, dxf=dxf)
        ws.conditional_formatting.add(f'A1:A{df.shape[0]+1}', rule)
        ws.conditional_formatting.add(f'B1:B{df.shape[0]+1}', rule) if not oneport else None


    def add_data_to_sheet(self, ws, df, oneport):
        ws['E2'] = 'Lower limit'
        ws['E2'].alignment = Alignment(horizontal='right', vertical='center', wrapText=False)
        ws['E3'] = 'Upper limit'
        ws['E3'].alignment = Alignment(horizontal='right', vertical='center', wrapText=False)

        self.num_format(ws['F2'])
        ws['F2'] = self.infusion_lower_limit

        self.num_format(ws['F3'])
        ws['F3'] = self.infusion_upper_limit

        self.num_format(ws['G2']) if not oneport else None
        ws['G2'] = self.injection_lower_limit if not oneport else None

        self.num_format(ws['G3'])
        ws['G3'] = self.injection_upper_limit if not oneport else None

        ws['E5'] = 'Average'
        ws['E5'].alignment = Alignment(horizontal='right', vertical='center', wrapText=False)
        ws['E6'] = 'Target'
        ws['E6'].alignment = Alignment(horizontal='right', vertical='center', wrapText=False)
        ws['E7'] = 'Delta'
        ws['E7'].alignment = Alignment(horizontal='right', vertical='center', wrapText=False)
        ws['E9'] = 'Installed shim'
        ws['E9'].alignment = Alignment(horizontal='right', vertical='center', wrapText=False)
        ws['E10'] = 'Required shim'
        ws['E10'].alignment = Alignment(horizontal='right', vertical='center', wrapText=False)

        ws['F1'].font = Font(bold=True)
        ws['F1'] = 'Infusion evaluation'

        value_false = "FALSCH" if self.language == 'de_DE' else "FALSE"
        
        self.num_format(ws['F5'])
        ws['F5'] = f'=AVERAGEIF(C2:C{df.shape[0]+1}, "{value_false}", A2:A{df.shape[0]+1})' # Infusion Mittelwert
        
        self.num_format(ws['F6'])
        ws['F6'] = '0,65' if self.language == 'de_DE' else '0.65' # Infusion Optimal
        
        self.num_format(ws['F7'])
        ws['F7'].font = Font(bold=True)
        ws['F7'] = '=F6-F5' # Infusion Optimal

        ws['G1'].font = Font(bold=True) if not oneport else None
        ws['G1'] = 'Injection evaluation' if not oneport else None
        
        self.num_format(ws['G5']) if not oneport else None
        ws['G5'] = f'=AVERAGEIF(D2:D{df.shape[0]+1}, "{value_false}", B2:B{df.shape[0]+1})' if not oneport else None # Injection Mittelwert
        
        self.num_format(ws['G6']) if not oneport else None
        ws['G6'] = ('0,40' if self.language == 'de_DE' else '0.40') if not oneport else None # Injection Optimal
        
        self.num_format(ws['G7']) if not oneport else None
        ws['G7'].font = Font(bold=True) if not oneport else None
        ws['G7'] = '=G6-G5' if not oneport else None # Injection Optimal
        
        # Legend
        ws['I16'].font = Font(bold=True)
        ws['I16'] = 'Legend'

        ws['I17'] = 'Smallest value'
        ws['J17'] = '0.001'
        ws['J17'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        ws['I18'] = 'Largest value'
        ws['J18'] = '0.7'
        ws['J18'].fill = PatternFill(start_color='B8F589', end_color='B8F589', fill_type='solid')


        ws['I19'] = 'Lower limit'
        ws['J19'].font = Font(bold=True, color='E67E17')
        ws['J19'] = '0.001'

        ws['I20'] = 'Upper limit'
        ws['J20'].font = Font(bold=True, color='D9112A')
        ws['J20'] = '1.0'